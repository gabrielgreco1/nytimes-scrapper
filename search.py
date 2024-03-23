from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import pandas as pd
import os
from openpyxl import load_workbook

class NYTSearch:
    def __init__(self, query):
        self.query = query
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        self.wait = WebDriverWait(self.driver, 10)

    def open_search(self):
        self.driver.get(f"https://www.nytimes.com/search?dropmab=false&query={self.query}&sort=newest")



    def scrape_list_items(self):
        try:
            # Wait until the list appears on the webpage
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'ol')))
            # Fetch all list items
            list_items = self.driver.find_elements(By.CSS_SELECTOR, 'ol > li')
            return list_items
        except TimeoutError:
            print("The list did not load in time")
            return []

    def scrape_item_details(self, list_item):
        # Assuming list_item is a WebElement object
        details = {}

        # Extract <p> elements
        details['paragraphs'] = [p.text for p in list_item.find_elements(By.CSS_SELECTOR, 'p')]

        # Extract <span> element
        details['date'] = [span.text for span in list_item.find_elements(By.CSS_SELECTOR, "span")]

        # Extract <a> elements
        details['links'] = [a.get_attribute('href') for a in list_item.find_elements(By.CSS_SELECTOR, 'a')]

        # Extract headings
        details['headings'] = [heading.text for heading in list_item.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6')]

        # Extract text from the span if present
        span_element = list_item.find_element(By.XPATH, '//*[@id="site-content"]/div/div[1]')
        figures = list_item.find_elements(By.CSS_SELECTOR, 'div > figure')
        images = [fig.find_element(By.TAG_NAME, 'img').get_attribute('src') for fig in figures if fig.find_element(By.TAG_NAME, 'img')]

        details['date_text'] = span_element.text if span_element else None

        return {
            'date_text': details,
            'image_urls': images
        }
    
    def download_image(self, url, path):
        response = requests.get(url)
        if response.status_code == 200:
            with open(path, 'wb') as file:
                file.write(response.content)
            print(f"Imagem baixada com sucesso: {path}")
        else:
            print(f"Erro ao baixar a imagem. Status code: {response.status_code}")


    def save_to_xlsx(self, data, file_path='C:\\Users\\gabri\\Desktop\\Challenge_RPA_PixelDu\\scrapped_items\\data.xlsx'):
        # Converte os dados para um DataFrame
        df = pd.DataFrame([data])
        
        # Verifica se o arquivo já existe
        if os.path.exists(file_path):
            book = load_workbook(file_path)
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets['Sheet1'].max_row
            
            # Se o arquivo existe, adiciona os novos dados na próxima linha vazia
            df.to_excel(writer, startrow=startrow, index=False, header=False, sheet_name='Sheet1')
        else:
            # Se o arquivo não existe, cria um novo e usa o index do DataFrame como identificador
            df.to_excel(file_path, index=False)

        writer.save()
        writer.close()

    def run(self):
        # Open the search page
        self.open_search()

        # Scrape all list items
        list_items = self.scrape_list_items()
        all_item_details = []

        for item in list_items:
            # Scrape details for each item
            details = self.scrape_item_details(item)
            if details:
                # print(f'News {item}')
                print(details)
                # print(f"Data: {details['date_text']}")
                # print(f"URLs das Imagens: {details['image_urls']}")
                print('-------------------------------------------------------------')

                for url in details['image_urls']:
                        filename = f"{details['date_text']['headings']}.jpg".replace('[', '').replace(']', '')
                        print(filename)
                        save_path = f"C:\\Users\\gabri\\Desktop\\Challenge_RPA_PixelDu\\images\\{filename}"
                        self.download_image(url, save_path)
               


        self.driver.quit()
        return all_item_details